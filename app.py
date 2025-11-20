import streamlit as st
import openpyxl
from deep_translator import GoogleTranslator
import io
import os
from concurrent.futures import ThreadPoolExecutor

# --- FIKSNI REČNIK (Sigurnosni filter) ---
# Ovde ubacujemo reči koje Google pogrešno prevodi.
# Skripta prvo gleda ovde, pa tek onda pita Google.
CUSTOM_DICTIONARY = {
    "Injection/Brizganje": "ইনজেকশন (Injection)",
    "Injection": "ইনজেকশন",
    "Brizganje": "ইনজেকশন (Injection)",
    "Squirting": "Injection", # Za svaki slučaj :)
    # Možeš dodati još termina ovde po potrebi format: "Srpski": "Bengalski"
}

# --- KONFIGURACIJA ---
st.set_page_config(page_title="Brzi Excel Prevodilac", layout="centered")

st.title("🇧🇩 Excel Prevodilac (Safe & Fast)")
st.markdown("""
Ova verzija:
1. Koristi **Fiksni rečnik** za tehničke termine (nema "čudnih" prevoda).
2. Koristi **Paralelno procesiranje** za veću brzinu.
3. Izbacuje čist sheet bez formula.
""")

# --- FUNKCIJA ZA PARALELNI PREVOD ---
def translate_text_worker(text):
    """Ova funkcija prevodi jednu reč. Koristi se u threadovima."""
    text = text.strip()
    
    # 1. Provera u fiksnom rečniku (Override)
    # Proveravamo da li se fraza sadrži u našem rečniku
    for key, value in CUSTOM_DICTIONARY.items():
        if key.lower() in text.lower():
            return value
            
    # 2. Ako nije u rečniku, pitaj Google
    try:
        # Instanciramo prevodioca ovde da ne bi došlo do konflikta threadova
        translator = GoogleTranslator(source='sr', target='bn')
        return translator.translate(text)
    except:
        return text # Ako pukne, vrati original

def translate_single_sheet_optimized(file, sheet_name):
    # 1. Učitavanje
    wb = openpyxl.load_workbook(file, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        return None
        
    ws = wb[sheet_name]
    ws.title = "Bengali_Prevod"
    
    # 2. Brisanje ostalih
    for name in wb.sheetnames:
        if name != "Bengali_Prevod":
            del wb[name]
            
    # 3. Skeniranje unikatnog teksta
    unique_texts = set()
    cells_map = [] 

    progress_bar = st.progress(0, text="Analiziram sadržaj...")
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip()
                if val and not val.isdigit():
                    unique_texts.add(val)
                    cells_map.append(cell)

    total_unique = len(unique_texts)
    st.caption(f"Broj unikatnih fraza za prevod: {total_unique}")
    
    if total_unique > 0:
        unique_list = list(unique_texts)
        translation_dict = {}
        
        progress_bar.progress(10, text=f"Pokrećem Turbo Prevod ({total_unique} fraza)...")
        
        # --- TURBO MODE (ThreadPoolExecutor) ---
        # Umesto jednog po jednog, šaljemo 10 zahteva odjednom
        with ThreadPoolExecutor(max_workers=10) as executor:
            results = list(executor.map(translate_text_worker, unique_list))
            
        # Spajamo rezultate
        for original, translated in zip(unique_list, results):
            translation_dict[original] = translated
            
        # 5. Primena prevoda
        progress_bar.progress(90, text="Upisujem podatke...")
        for cell in cells_map:
            original = cell.value.strip()
            if original in translation_dict:
                cell.value = translation_dict[original]

    # 6. Čuvanje
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    progress_bar.progress(100, text="Gotovo!")
    
    return output

# --- INTERFEJS ---
uploaded_file = st.file_uploader("Izaberi Excel fajl", type=["xlsx"])

if uploaded_file is not None:
    try:
        wb_temp = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet_names = wb_temp.sheetnames
        wb_temp.close()
        
        selected_sheet = st.selectbox("Koji sheet prevodimo?", sheet_names)
        
        if st.button("🚀 Prevedi"):
            with st.spinner('Radim...'):
                result = translate_single_sheet_optimized(uploaded_file, selected_sheet)
                
                if result:
                    st.success("Završeno!")
                    
                    # --- KREIRANJE IMENA FAJLA ---
                    # Uzimamo originalno ime bez ekstenzije
                    original_name = os.path.splitext(uploaded_file.name)[0]
                    # Format: ImeFajla_ImeSheeta_BN.xlsx
                    new_filename = f"{original_name}_{selected_sheet}_BN.xlsx"
                    
                    st.download_button(
                        label="📥 Preuzmi Prevedeni Fajl",
                        data=result,
                        file_name=new_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.error("Došlo je do greške.")
                    
    except Exception as e:
        st.error(f"Greška: {e}")
